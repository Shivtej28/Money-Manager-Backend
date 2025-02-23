
from datetime import datetime
from urllib import response
from fastapi import APIRouter, Depends, UploadFile, HTTPException
from ..dao.database import get_db
from sqlalchemy.orm import Session
from ..utils.jwtauth import decode_jwt_token
from ..scehmas.request_model import TransactionBase, TransactionResponse, Response, UploadTransaction
from ..service.transactioin_service import transaction_service
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
from typing import List
from io import BytesIO
from openpyxl import load_workbook

TOKEN = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOjIsImV4cCI6MTcyNzUwODkyOH0.bloAuZfhrNj_ODh5vIFqUsJXEYTUUbfmqYFOG0BCI14"
router = APIRouter(prefix="/api/transactions", tags=['Transactions'])


@router.get("/", response_model=Response[List[TransactionResponse]])
def get_all_transactions(db: Session = Depends(get_db), user=Depends(decode_jwt_token)):
    response = transaction_service.get_all_transactions(db, user)

    return JSONResponse(status_code=response.status_code, content=jsonable_encoder(response))


@router.post("/", response_model=Response)
def create_transaction(transaction: TransactionBase, db: Session = Depends(get_db), user=Depends(decode_jwt_token)):
    add_transaction = transaction.dict()
    response = transaction_service.create_transaction(
        add_transaction, db, user)
    print("IN API_--------------------")
    # print(response)
    return JSONResponse(status_code=response.status_code, content=jsonable_encoder(response))


@router.put("/{id}", response_model=Response)
def update_transaction(id: int, transaction: TransactionBase, db: Session = Depends(get_db), user=Depends(decode_jwt_token)):
    response = transaction_service.update_transaction(
        id, transaction, db, user)
    return JSONResponse(status_code=response.status_code, content=jsonable_encoder(response))


@router.delete("/{id}")
def delete_transaction(id: int, db: Session = Depends(get_db), user=Depends(decode_jwt_token)):
    response = transaction_service.delete_transaction(id, db, user)
    return JSONResponse(status_code=response.status_code, content=jsonable_encoder(response))


@router.post("/upload-statement")
async def upload_statment(file: UploadFile, db: Session = Depends(get_db), user=Depends(decode_jwt_token)):
    if not (file.filename.endswith(".xls") or file.filename.endswith(".xlsx")):
        raise HTTPException(
            status_code=400, detail="Invalid file format. Only .xls and .xlsx files are allowed.")

    print("File Name--------", file.filename)
    content = await file.read()
    print(file.content_type)
    workook = load_workbook(BytesIO(content))
    sheet = workook.active

    transactions = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping header row
        date, description, ref_no, value, withdraw_amount, deposit_amount, balance = row
        t = UploadTransaction(amount=deposit_amount if withdraw_amount is None else withdraw_amount,
                              transaction_date=datetime.strptime(
                                  date, "%d/%m/%y").date(),
                              transaction_type="Income" if withdraw_amount is None else "Expense",
                              bank_id=1,
                              description=description,
                              category_id=1,
                              subcategory_id=1)
        transactions.append(t)
    response = transaction_service.upload_file(transactions, db, user)

    return response
